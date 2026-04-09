"""
Export Module — Generate PDF and Excel reports from analysis data.
"""

import io
import json
import os
from datetime import datetime
from typing import Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

from fpdf import FPDF


# ═══════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════

def generate_excel(statement: dict, analysis: dict) -> io.BytesIO:
    """Generate a comprehensive Excel report with multiple sheets."""
    wb = Workbook()

    # Colors
    header_fill = PatternFill(start_color="1A2B50", end_color="1A2B50", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1A2B50")
    sub_font = Font(name="Calibri", size=10, color="666666")
    money_font = Font(name="Calibri", size=11, color="1A2B50")
    credit_font = Font(name="Calibri", size=11, color="16A34A", bold=True)
    debit_font = Font(name="Calibri", size=11, color="DC2626", bold=True)
    thin_border = Border(
        bottom=Side(style="thin", color="E5E7EB")
    )

    ai = statement.get("account_info", {})
    txns = statement.get("transactions", [])
    sc = analysis.get("summary_card", {})
    hs = analysis.get("health_score", {})

    # ─── Sheet 1: Summary ───
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1A2B50"

    ws.merge_cells("A1:F1")
    ws["A1"] = "MakeEazy Bank Statement Analysis Report"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color="1A2B50")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ws["A2"].font = sub_font

    # Account Info
    info_rows = [
        ("Account Holder", ai.get("account_holder_name", "")),
        ("Bank", ai.get("bank_name", "").upper()),
        ("Account Number", ai.get("account_number", "")),
        ("IFSC Code", ai.get("ifsc", "")),
        ("Period", f"{ai.get('statement_period', {}).get('from', '')} to {ai.get('statement_period', {}).get('to', '')}"),
        ("Health Score", f"{hs.get('score', 0)}/100 ({hs.get('rating', '')})"),
    ]

    ws["A4"] = "ACCOUNT PROFILE"
    ws["A4"].font = Font(name="Calibri", size=12, bold=True, color="E8722A")

    for i, (label, value) in enumerate(info_rows, start=5):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(name="Calibri", size=10, bold=True, color="666666")
        ws[f"B{i}"] = str(value)
        ws[f"B{i}"].font = Font(name="Calibri", size=11, color="1A2B50")

    # Financial Summary
    bs = sc.get("balance_summary", {})
    ts_data = sc.get("transaction_summary", {})
    fs = sc.get("financial_summary", {})

    ws["A13"] = "FINANCIAL SUMMARY"
    ws["A13"].font = Font(name="Calibri", size=12, bold=True, color="E8722A")

    fin_rows = [
        ("Total Transactions", ts_data.get("total", 0)),
        ("Total Credits", ts_data.get("total_credits", 0)),
        ("Total Debits", ts_data.get("total_debits", 0)),
        ("Credit Count", ts_data.get("credit_count", 0)),
        ("Debit Count", ts_data.get("debit_count", 0)),
        ("Net Cash Flow", ts_data.get("net_flow", 0)),
        ("Opening Balance", bs.get("opening_balance", 0)),
        ("Closing Balance", bs.get("closing_balance", 0)),
        ("Average Balance", bs.get("average_balance", 0)),
        ("Max Balance", bs.get("max_balance", 0)),
        ("Min Balance", bs.get("min_balance", 0)),
    ]

    for i, (label, value) in enumerate(fin_rows, start=14):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(name="Calibri", size=10, bold=True, color="666666")
        cell = ws[f"B{i}"]
        if isinstance(value, (int, float)):
            cell.value = value
            cell.number_format = '#,##0.00'
        else:
            cell.value = str(value)
        cell.font = money_font

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 25

    # ─── Sheet 2: Transactions ───
    ws2 = wb.create_sheet("Transactions")
    ws2.sheet_properties.tabColor = "16A34A"

    headers = ["Sr No", "Date", "Description", "Debit", "Credit", "Balance", "Mode", "Category"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, t in enumerate(txns, start=2):
        ws2.cell(row=row_idx, column=1, value=t.get("sr_no", row_idx - 1))
        ws2.cell(row=row_idx, column=2, value=t.get("txn_date", ""))
        ws2.cell(row=row_idx, column=3, value=t.get("description", ""))

        debit_cell = ws2.cell(row=row_idx, column=4)
        debit_val = t.get("debit", 0)
        if debit_val:
            debit_cell.value = debit_val
            debit_cell.number_format = '#,##0.00'
            debit_cell.font = debit_font

        credit_cell = ws2.cell(row=row_idx, column=5)
        credit_val = t.get("credit", 0)
        if credit_val:
            credit_cell.value = credit_val
            credit_cell.number_format = '#,##0.00'
            credit_cell.font = credit_font

        bal_cell = ws2.cell(row=row_idx, column=6, value=t.get("balance", 0))
        bal_cell.number_format = '#,##0.00'
        bal_cell.font = money_font

        ws2.cell(row=row_idx, column=7, value=t.get("txn_mode", ""))
        ws2.cell(row=row_idx, column=8, value=t.get("ai_category", ""))

        for col in range(1, 9):
            ws2.cell(row=row_idx, column=col).border = thin_border

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 50
    ws2.column_dimensions["D"].width = 15
    ws2.column_dimensions["E"].width = 15
    ws2.column_dimensions["F"].width = 15
    ws2.column_dimensions["G"].width = 12
    ws2.column_dimensions["H"].width = 14

    ws2.auto_filter.ref = f"A1:H{len(txns) + 1}"

    # ─── Sheet 3: Monthly Summary ───
    ms = analysis.get("monthly_summary", {}).get("months", [])
    if ms:
        ws3 = wb.create_sheet("Monthly Summary")
        ws3.sheet_properties.tabColor = "E8722A"
        m_headers = ["Month", "Credits", "Debits", "Net Flow", "Txn Count"]
        for col, h in enumerate(m_headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, m in enumerate(ms, start=2):
            ws3.cell(row=i, column=1, value=m.get("month", ""))
            c = ws3.cell(row=i, column=2, value=m.get("total_credit", 0))
            c.number_format = '#,##0.00'
            c.font = credit_font
            d = ws3.cell(row=i, column=3, value=m.get("total_debit", 0))
            d.number_format = '#,##0.00'
            d.font = debit_font
            n = ws3.cell(row=i, column=4, value=m.get("net_flow", 0))
            n.number_format = '#,##0.00'
            ws3.cell(row=i, column=5, value=m.get("transaction_count", 0))

        for col_letter in ["A", "B", "C", "D", "E"]:
            ws3.column_dimensions[col_letter].width = 18

    # ─── Sheet 4: Salary Analysis ───
    sal = analysis.get("salary_analysis", {})
    sal_txns = sal.get("salary_transactions", [])
    if sal_txns:
        ws4 = wb.create_sheet("Salary")
        ws4.sheet_properties.tabColor = "2563EB"
        s_headers = ["Date", "Description", "Amount", "Source"]
        for col, h in enumerate(s_headers, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for i, s in enumerate(sal_txns, start=2):
            ws4.cell(row=i, column=1, value=s.get("date", ""))
            ws4.cell(row=i, column=2, value=s.get("description", ""))
            c = ws4.cell(row=i, column=3, value=s.get("amount", 0))
            c.number_format = '#,##0.00'
            c.font = credit_font
            ws4.cell(row=i, column=4, value=s.get("source", ""))
        for col_letter in ["A", "B", "C", "D"]:
            ws4.column_dimensions[col_letter].width = 22

    # ─── Sheet 5: EMI / Obligations ───
    emi = analysis.get("emi_obligations", {})
    emi_txns = emi.get("emi_transactions", [])
    if emi_txns:
        ws5 = wb.create_sheet("EMI & Obligations")
        ws5.sheet_properties.tabColor = "DC2626"
        e_headers = ["Date", "Description", "Amount"]
        for col, h in enumerate(e_headers, 1):
            cell = ws5.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for i, e in enumerate(emi_txns, start=2):
            ws5.cell(row=i, column=1, value=e.get("date", ""))
            ws5.cell(row=i, column=2, value=e.get("description", ""))
            c = ws5.cell(row=i, column=3, value=e.get("amount", 0))
            c.number_format = '#,##0.00'
            c.font = debit_font
        for col_letter in ["A", "B", "C"]:
            ws5.column_dimensions[col_letter].width = 25

    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ═══════════════════════════════════════════════════════
# PDF EXPORT
# ═══════════════════════════════════════════════════════

class PDFReport(FPDF):
    """Custom PDF report with MakeEazy branding."""

    def header(self):
        self.set_fill_color(26, 43, 80)
        self.rect(0, 0, 210, 18, 'F')
        self.set_text_color(255, 255, 255)
        self.set_font("Helvetica", "B", 12)
        self.cell(0, 12, "MakeEazy Bank Statement Analysis", align="L")
        self.set_font("Helvetica", "", 8)
        self.cell(0, 12, f"Generated: {datetime.now().strftime('%d-%b-%Y')}", align="R")
        self.ln(18)

    def footer(self):
        self.set_y(-10)
        self.set_font("Helvetica", "I", 7)
        self.set_text_color(150, 150, 150)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}} | MakeEazy Bank Analyzer v1.0", align="C")

    def section_title(self, title):
        self.set_font("Helvetica", "B", 11)
        self.set_text_color(232, 114, 42)
        self.cell(0, 8, title, ln=True)
        self.set_draw_color(232, 114, 42)
        self.line(self.get_x(), self.get_y(), self.get_x() + 190, self.get_y())
        self.ln(3)

    def info_row(self, label, value):
        self.set_font("Helvetica", "", 9)
        self.set_text_color(100, 100, 100)
        self.cell(50, 6, label, align="L")
        self.set_font("Helvetica", "B", 9)
        self.set_text_color(26, 43, 80)
        self.cell(0, 6, str(value), ln=True)

    def money(self, amount):
        try:
            return f"Rs.{float(amount):,.2f}"
        except (ValueError, TypeError):
            return str(amount)


def generate_pdf(statement: dict, analysis: dict) -> io.BytesIO:
    """Generate a comprehensive PDF report."""
    pdf = PDFReport()
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    ai = statement.get("account_info", {})
    txns = statement.get("transactions", [])
    sc = analysis.get("summary_card", {})
    bs = sc.get("balance_summary", {})
    ts = sc.get("transaction_summary", {})
    hs = analysis.get("health_score", {})

    # ─── Account Profile ───
    pdf.section_title("ACCOUNT PROFILE")
    pdf.info_row("Account Holder", ai.get("account_holder_name", ""))
    pdf.info_row("Bank", ai.get("bank_name", "").upper())
    pdf.info_row("Account Number", ai.get("account_number", ""))
    pdf.info_row("IFSC Code", ai.get("ifsc", ""))
    period = ai.get("statement_period", {})
    pdf.info_row("Period", f"{period.get('from', '')} to {period.get('to', '')}")
    pdf.info_row("Health Score", f"{hs.get('score', 0)}/100 ({hs.get('grade', '')})")
    pdf.ln(6)

    # ─── Financial Summary ───
    pdf.section_title("FINANCIAL SUMMARY")
    pdf.info_row("Total Transactions", str(ts.get("total", 0)))
    pdf.info_row("Total Credits", pdf.money(ts.get("total_credits", 0)))
    pdf.info_row("Total Debits", pdf.money(ts.get("total_debits", 0)))
    pdf.info_row("Net Cash Flow", pdf.money(ts.get("net_flow", 0)))
    pdf.info_row("Opening Balance", pdf.money(bs.get("opening_balance", 0)))
    pdf.info_row("Closing Balance", pdf.money(bs.get("closing_balance", 0)))
    pdf.info_row("Average Balance", pdf.money(bs.get("average_balance", 0)))
    pdf.ln(6)

    # ─── Average Bank Balance ───
    abb = analysis.get("average_bank_balance", {})
    if abb:
        pdf.section_title("AVERAGE BANK BALANCE")
        pdf.info_row("ABB (Last 30 Days)", pdf.money(abb.get("abb_last_30_days", 0)))
        pdf.info_row("ABB (Last 3 Months)", pdf.money(abb.get("abb_last_3_months", 0)))
        pdf.info_row("ABB (Last 6 Months)", pdf.money(abb.get("abb_last_6_months", 0)))
        pdf.ln(6)

    # ─── Monthly Summary Table ───
    ms = analysis.get("monthly_summary", {}).get("months", [])
    if ms:
        pdf.section_title("MONTHLY SUMMARY")
        # Table header
        pdf.set_fill_color(26, 43, 80)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 8)
        col_widths = [30, 35, 35, 35, 25, 30]
        headers = ["Month", "Credits", "Debits", "Net Flow", "Txns", "Avg Bal"]
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], 7, h, border=1, align="C", fill=True)
        pdf.ln()

        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(50, 50, 50)
        for m in ms:
            pdf.cell(col_widths[0], 6, m.get("month", ""), border=1, align="C")
            pdf.cell(col_widths[1], 6, pdf.money(m.get("total_credit", 0)), border=1, align="R")
            pdf.cell(col_widths[2], 6, pdf.money(m.get("total_debit", 0)), border=1, align="R")
            nf = m.get("net_flow", 0)
            pdf.cell(col_widths[3], 6, pdf.money(nf), border=1, align="R")
            pdf.cell(col_widths[4], 6, str(m.get("transaction_count", 0)), border=1, align="C")
            pdf.cell(col_widths[5], 6, pdf.money(m.get("avg_balance", 0)), border=1, align="R")
            pdf.ln()
        pdf.ln(6)

    # ─── Salary Analysis ───
    sal = analysis.get("salary_analysis", {})
    sal_txns = sal.get("salary_transactions", [])
    if sal_txns:
        pdf.add_page()
        pdf.section_title("SALARY ANALYSIS")
        pdf.info_row("Salary Detected", "Yes" if sal.get("salary_detected") else "No")
        pdf.info_row("Avg Monthly Salary", pdf.money(sal.get("average_monthly_salary", 0)))
        pdf.info_row("Total Salary Income", pdf.money(sal.get("total_salary_amount", 0)))
        pdf.ln(4)

        pdf.set_fill_color(26, 43, 80)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 8)
        for w, h in [(25, "Date"), (95, "Description"), (30, "Amount"), (40, "Source")]:
            pdf.cell(w, 7, h, border=1, align="C", fill=True)
        pdf.ln()

        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(50, 50, 50)
        for s in sal_txns[:30]:
            pdf.cell(25, 6, s.get("date", ""), border=1)
            desc = s.get("description", "")[:55]
            pdf.cell(95, 6, desc, border=1)
            pdf.cell(30, 6, pdf.money(s.get("amount", 0)), border=1, align="R")
            pdf.cell(40, 6, s.get("source", "")[:25], border=1)
            pdf.ln()
        pdf.ln(6)

    # ─── Top Transactions ───
    top = analysis.get("top_transactions", {})
    top_debits = top.get("top_debits", [])[:10]
    top_credits = top.get("top_credits", [])[:10]

    if top_debits or top_credits:
        pdf.add_page()
        pdf.section_title("TOP TRANSACTIONS")

        if top_credits:
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_text_color(22, 163, 74)
            pdf.cell(0, 7, "Top Credits", ln=True)
            pdf.set_fill_color(26, 43, 80)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 8)
            for w, h in [(25, "Date"), (110, "Description"), (30, "Amount")]:
                pdf.cell(w, 7, h, border=1, align="C", fill=True)
            pdf.ln()
            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(50, 50, 50)
            for t in top_credits:
                pdf.cell(25, 6, t.get("date", ""), border=1)
                pdf.cell(110, 6, t.get("description", "")[:65], border=1)
                pdf.cell(30, 6, pdf.money(t.get("amount", 0)), border=1, align="R")
                pdf.ln()
            pdf.ln(4)

        if top_debits:
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_text_color(220, 38, 38)
            pdf.cell(0, 7, "Top Debits", ln=True)
            pdf.set_fill_color(26, 43, 80)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 8)
            for w, h in [(25, "Date"), (110, "Description"), (30, "Amount")]:
                pdf.cell(w, 7, h, border=1, align="C", fill=True)
            pdf.ln()
            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(50, 50, 50)
            for t in top_debits:
                pdf.cell(25, 6, t.get("date", ""), border=1)
                pdf.cell(110, 6, t.get("description", "")[:65], border=1)
                pdf.cell(30, 6, pdf.money(t.get("amount", 0)), border=1, align="R")
                pdf.ln()

    # ─── Transactions (last pages) ───
    pdf.add_page()
    pdf.section_title(f"ALL TRANSACTIONS ({len(txns)} entries)")

    pdf.set_fill_color(26, 43, 80)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 7)
    cols = [(8, "#"), (18, "Date"), (80, "Description"), (22, "Debit"), (22, "Credit"), (22, "Balance"), (18, "Mode")]
    for w, h in cols:
        pdf.cell(w, 6, h, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(50, 50, 50)
    for t in txns:
        if pdf.get_y() > 270:
            pdf.add_page()
            pdf.set_fill_color(26, 43, 80)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 7)
            for w, h in cols:
                pdf.cell(w, 6, h, border=1, align="C", fill=True)
            pdf.ln()
            pdf.set_font("Helvetica", "", 7)
            pdf.set_text_color(50, 50, 50)

        pdf.cell(8, 5, str(t.get("sr_no", "")), border=1, align="C")
        pdf.cell(18, 5, t.get("txn_date", ""), border=1)
        pdf.cell(80, 5, t.get("description", "")[:50], border=1)
        d = t.get("debit", 0)
        c = t.get("credit", 0)
        pdf.cell(22, 5, f"{d:,.2f}" if d else "", border=1, align="R")
        pdf.cell(22, 5, f"{c:,.2f}" if c else "", border=1, align="R")
        pdf.cell(22, 5, f"{t.get('balance', 0):,.2f}", border=1, align="R")
        pdf.cell(18, 5, t.get("txn_mode", "")[:8], border=1, align="C")
        pdf.ln()

    buffer = io.BytesIO()
    pdf.output(buffer)
    buffer.seek(0)
    return buffer
