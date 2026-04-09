"""
Excel Parser — Extracts transactions from bank statement Excel files (.xlsx, .xls).
Uses openpyxl directly (no pandas dependency for low memory).
"""

import re
from typing import Dict, Any, List, Optional
from datetime import datetime

from app.services.parser.bank_detector import detect_bank_from_text


COLUMN_MAPPINGS = {
    "date": ["date", "txn date", "transaction date", "trans date", "posting date", "value date",
             "txn dt", "trans dt", "dated"],
    "value_date": ["value date", "val date", "value dt"],
    "description": ["description", "narration", "particulars", "details", "remarks",
                     "transaction details", "transaction description", "trans description", "narrative"],
    "reference": ["reference", "ref no", "ref", "chq no", "cheque no", "utr", "txn ref",
                   "reference no", "instrument no"],
    "debit": ["debit", "withdrawal", "dr", "debit amount", "withdrawals", "debit(dr)",
              "amount(dr)", "dr amount", "debit (inr)"],
    "credit": ["credit", "deposit", "cr", "credit amount", "deposits", "credit(cr)",
               "amount(cr)", "cr amount", "credit (inr)"],
    "balance": ["balance", "closing balance", "running balance", "available balance",
                "bal", "closing bal", "balance (inr)"],
}


def _map_columns(headers: List[str]) -> Dict[str, int]:
    """Map actual column names to standard field names."""
    col_map = {}
    for idx, header in enumerate(headers):
        if not header:
            continue
        header_lower = str(header).strip().lower()
        for field, keywords in COLUMN_MAPPINGS.items():
            if any(kw == header_lower or kw in header_lower for kw in keywords):
                if field not in col_map:
                    col_map[field] = idx
                    break
    return col_map


def _normalize_date(date_val) -> str:
    """Normalize date values from Excel."""
    if date_val is None:
        return ""
    if isinstance(date_val, datetime):
        return date_val.strftime("%d-%m-%y")
    date_str = str(date_val).strip()
    if not date_str:
        return ""
    formats = [
        "%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y", "%d/%m/%y",
        "%Y-%m-%d", "%d %b %Y", "%d %b %y", "%d-%b-%Y", "%d-%b-%y",
        "%m/%d/%Y", "%d.%m.%Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).strftime("%d-%m-%y")
        except ValueError:
            continue
    return date_str


def _parse_amount(val) -> float:
    """Parse amount value."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return abs(val)
    cleaned = re.sub(r'[₹,\s"\'()]', '', str(val).strip())
    cleaned = re.sub(r'(Dr\.?|Cr\.?|DR|CR)$', '', cleaned, flags=re.IGNORECASE).strip()
    try:
        return abs(float(cleaned))
    except (ValueError, TypeError):
        return 0.0


def parse_excel(file_path: str) -> Dict[str, Any]:
    """Parse Excel bank statement file using openpyxl."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # Read all rows
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))
    wb.close()

    if not all_rows:
        raise ValueError("Excel file is empty.")

    # Find header row
    header_idx = 0
    for idx in range(min(20, len(all_rows))):
        row_text = ' '.join(str(v).lower() for v in all_rows[idx] if v is not None)
        if any(kw in row_text for kw in ["date", "narration", "description", "particulars"]):
            if any(kw in row_text for kw in ["debit", "credit", "withdrawal", "deposit", "balance", "amount"]):
                header_idx = idx
                break

    headers = [str(v).strip() if v else "" for v in all_rows[header_idx]]
    data_rows = all_rows[header_idx + 1:]

    # Detect bank
    all_text = ' '.join([' '.join(str(v) for v in row if v) for row in all_rows[:30]])
    bank_key, _ = detect_bank_from_text(all_text[:2000])

    # Map columns
    col_map = _map_columns(headers)

    if "date" not in col_map:
        raise ValueError("Could not identify date column in Excel file.")

    # Extract transactions
    transactions = []
    for row in data_rows:
        date_idx = col_map["date"]
        if date_idx >= len(row) or row[date_idx] is None:
            continue
        date_str = _normalize_date(row[date_idx])
        if not date_str:
            continue

        def get_val(field):
            idx = col_map.get(field)
            if idx is not None and idx < len(row):
                return row[idx]
            return None

        debit = _parse_amount(get_val("debit"))
        credit = _parse_amount(get_val("credit"))

        txn = {
            "sr_no": len(transactions) + 1,
            "txn_date": date_str,
            "value_date": _normalize_date(get_val("value_date")) if "value_date" in col_map else "",
            "reference_no": str(get_val("reference") or "").strip(),
            "description": str(get_val("description") or "").strip(),
            "debit": debit,
            "credit": credit,
            "balance": _parse_amount(get_val("balance")),
            "txn_type": "Dr." if debit > 0 else ("Cr." if credit > 0 else ""),
        }
        transactions.append(txn)

    account_info = {
        "bank_name": bank_key,
        "account_holder_name": "",
        "account_number": "",
        "address": "",
        "ifsc": "",
        "micr_code": "",
        "customer_id": "",
        "email": "",
        "phone": "",
        "account_type": "",
        "account_open_date": "",
        "branch_name": "",
        "statement_period": {
            "from": transactions[0]["txn_date"] if transactions else "",
            "to": transactions[-1]["txn_date"] if transactions else "",
        },
    }

    return {
        "account_info": account_info,
        "transactions": transactions,
        "mismatched_sequence_date": [],
        "negative_balance": [],
        "discrepancies": {"balance_errors": [], "swapped_credit_debit_rows": [], "corrected_row_indices": []},
    }
